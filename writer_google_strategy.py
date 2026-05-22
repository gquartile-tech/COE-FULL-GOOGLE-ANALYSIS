"""
writer_google_strategy.py
Strategy pillar — reviewer-driven, no scoring thresholds.
"""
from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from reader_databricks_google import GoogleContext, clean_text
from config import STATUS_OK, STATUS_FLAG, STATUS_PARTIAL

def write_strategy_output(template_path, output_path, results, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb["Account Strategy_Analysis"]
    ws_ref  = wb["Account Strategy_Reference"]

    ws_main["A1"] = f"{ctx.hash_name} — Google Account Strategy Analysis"
    ws_main["B3"] = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}"
    if ctx.window_start and ctx.window_end and ctx.window_days:
        ws_main["B4"] = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
    if ctx.downloaded:
        ws_main["B5"] = ctx.downloaded
        ws_main["B5"].number_format = "yyyy-mm-dd hh:mm:ss"

    cid_to_row = {}
    for r in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f"B{r}"].value).upper()
        if cid:
            cid_to_row[cid] = r

    for cid, res in results.items():
        if cid not in cid_to_row:
            continue
        rr = cid_to_row[cid]
        ws_ref[f"D{rr}"] = res.status
        ws_ref[f"H{rr}"] = res.what
        ws_ref[f"I{rr}"] = res.why
        for cell in [f"H{rr}", f"I{rr}"]:
            ws_ref[cell].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"[writer_strategy] Saved: {output_path}")
