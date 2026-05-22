"""
writer_google_implementation.py
Writes Implementation pillar results to CoE_Google_Account_Implementation_Analysis_Templates.xlsm.
Tab names: Account Implement_Analysis, Account Implement_Reference
"""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from reader_databricks_google import GoogleContext, clean_text
from config_google_implementation import SCORING_EXCLUDED, PRIORITY_POINTS, IMPORTANCE
from config import STATUS_OK, STATUS_FLAG, STATUS_PARTIAL

def _safe_write(ws, cell_addr, value):
    """Write to a cell, automatically resolving merged cell top-left anchor."""
    from openpyxl.cell import MergedCell
    cell = ws[cell_addr]
    if isinstance(cell, MergedCell):
        # Find the top-left cell of the merge range that contains this cell
        for merge_range in ws.merged_cells.ranges:
            if cell_addr in merge_range:
                top_left = ws.cell(merge_range.min_row, merge_range.min_col)
                top_left.value = value
                return
        # Fallback: skip silently if no range found
        return
    cell.value = value




def _compute_score(results: dict) -> tuple:
    score = 100
    for cid, res in results.items():
        if cid in SCORING_EXCLUDED:
            continue
        if res.status in (STATUS_FLAG, STATUS_PARTIAL):
            imp = IMPORTANCE.get(cid, 5)
            pts = PRIORITY_POINTS.get(imp, -5)
            if res.status == STATUS_PARTIAL:
                pts = pts // 2
            score += pts
    score = max(0, min(100, score))
    grade = "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 60 else "D" if score >= 45 else "F"
    return score, grade


def write_implementation_output(
    template_path: str,
    output_path: str,
    results: dict,
    ctx: GoogleContext,
) -> None:
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb["Account Implement_Analysis"]
    ws_ref  = wb["Account Implement_Reference"]

    score, grade = _compute_score(results)

    _safe_write(ws_main, "A1", f"{ctx.hash_name} — Google Implementation Analysis")
    _safe_write(ws_main, "B3", f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, "B4", f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        ws_main["B5"] = ctx.downloaded
        ws_main["B5"].number_format = "yyyy-mm-dd hh:mm:ss"
    _safe_write(ws_main, "C8", score)
    _safe_write(ws_main, "D8", grade)

    cid_to_row = {}
    for r in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f"B{r}"].value).upper()
        if cid:
            cid_to_row[cid] = r

    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f"[writer_implementation] WARNING: {cid} not in reference tab — skipping.")
            continue
        rr = cid_to_row[cid]
        ws_ref[f"D{rr}"] = res.status
        ws_ref[f"H{rr}"] = res.what
        ws_ref[f"I{rr}"] = res.why
        for cell in [f"H{rr}", f"I{rr}"]:
            ws_ref[cell].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"[writer_implementation] Saved: {output_path} | Score: {score} | Grade: {grade}")
