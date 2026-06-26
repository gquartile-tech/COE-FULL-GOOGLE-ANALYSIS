"""
writer_google_scorecard.py

Populates the Google Scorecard template (Google_Scorecard_-_v1_2026.xlsx) from
the combined results of all five Google CoE agent pillars.

Column D (col 4) = TRUE / FALSE / "PARTIAL" per scorecard row.
Conversion logic:
  OK      → TRUE
  FLAG    → FALSE
  PARTIAL → "PARTIAL"

KPI cells (Account Health section, rows 3-13) are populated from GoogleContext.
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import STATUS_OK, STATUS_FLAG, STATUS_PARTIAL, ControlResult
from reader_databricks_google import GoogleContext, get_sheet, find_col, to_float


# ── Status → cell value ───────────────────────────────────────────────────────

def _val(res: ControlResult):
    if res.status == STATUS_OK:
        return True
    if res.status == STATUS_PARTIAL:
        return "PARTIAL"
    return False


# ── Row mapping: (pillar_results_key, control_id) → scorecard row number ─────
# Column D is the value column (index 4, Excel col D).
# Rows that have no direct agent control are left as-is (template default FALSE).

ROW_MAP: Dict[int, tuple] = {
    # ── Implementation ────────────────────────────────────────────────────────
    15: ("google_implementation", "I004"),   # Billing Status
    # 16 Policy Violations → I005 (manual OK) — leave as-is
    17: ("google_implementation", "I003"),   # Conversion Actions → tag proxy
    18: ("google_implementation", "I002"),   # GMC and GA4 linked
    # 19 Product Status → manual (H023) — leave
    # 20 Data Sources → I010 manual — leave
    # 21 Feed Source → M007 partial proxy
    21: ("google_mastery",        "M007"),
    22: ("google_implementation", "I009"),   # Shopify Connection

    # ── Mastery ───────────────────────────────────────────────────────────────
    23: ("google_mastery", "M001"),   # Meeting Frequency
    24: ("google_mastery", "M002"),   # Touchpoint frequency
    25: ("google_mastery", "M003"),   # Budgets and Goals
    # 26 Asset Approval Conversation → no direct control
    # 27 Internal Slack → no direct control
    28: ("google_mastery", "M005"),   # DPL
    29: ("google_mastery", "M006"),   # Custom Labels
    30: ("google_mastery", "M007"),   # Transformers
    31: ("google_mastery", "M008"),   # Price Competitiveness
    32: ("google_mastery", "M009"),   # Feed Based Inv. Filters
    # 33 Titles/Descriptions → M007 proxy (transformer output quality)
    34: ("google_mastery", "M010"),   # Product Types
    35: ("google_mastery", "M011"),   # Brands

    # ── Framework ────────────────────────────────────────────────────────────
    36: ("google_framework", "F003"),  # Auto-Apply Settings Disabled
    37: ("google_framework", "F022"),  # Ad extensions (sitelinks)
    # 38 Custom Segment → no direct control
    39: ("google_framework", "F006"),  # Final URL proxy (promotion end = closest)
    40: ("google_framework", "F005"),  # Location Targeting
    41: ("google_framework", "F001"),  # Naming Conventions
    42: ("google_framework", "F013"),  # Negative KW lists
    43: ("google_framework", "F006"),  # Promotion End dates
    44: ("google_strategy",  "S021"),  # Display/Demand Gen - Optimized Targeting
    45: ("google_framework", "F004"),  # Paid Search - Display Expansion Disabled
    46: ("google_framework", "F007"),  # PMax - Extension Settings (PMAX automation)
    47: ("google_framework", "F026"),  # PMax - Images (logos/assets)
    48: ("google_framework", "F021"),  # PMax - Placements (PMAX channel distribution)
    49: ("google_framework", "F035"),  # PMax - Search Themes
    # 50 Search - Ad Disapprovals → no direct control
    51: ("google_framework", "F027"),  # Search - Ad Strength
    52: ("google_framework", "F031"),  # Shopping - Campaign Priorities
    53: ("google_framework", "F034"),  # Shopping/PMax - Audiences
    # 54 Shopping/PMax - ID Funneling → no direct control

    # ── Strategy — Shopping/PMax Core ────────────────────────────────────────
    56: ("google_strategy", "S001"),   # Catchall/EE
    57: ("google_strategy", "S002"),   # TOP Products
    58: ("google_strategy", "S003"),   # Price Tier/Margins
    59: ("google_strategy", "S004"),   # Brand Campaigns
    60: ("google_strategy", "S005"),   # Shopping Suppression
    61: ("google_strategy", "S006"),   # Product Type Campaigns

    # ── Strategy — Paid Search Core ───────────────────────────────────────────
    62: ("google_strategy", "S010"),   # Trademark Exact / SKW
    63: ("google_strategy", "S011"),   # NB Search
    64: ("google_strategy", "S012"),   # NB DSA

    # ── Strategy — Shopping/PMax extended ────────────────────────────────────
    # 65 CWCD → no single control (structural concept)
    66: ("google_strategy", "S007"),   # Zombie
    67: ("google_strategy", "S008"),   # Remnant
    68: ("google_strategy", "S009"),   # Query-Based
    # 69 Push Campaign → no direct control
    70: ("google_framework", "F020"),  # Shopping - Price Competitiveness
    71: ("google_strategy", "S014"),   # Devices (device bid adjustments)
    # 72 PMax Asset-Only → no direct control
    # 73 NCA → no direct control
    # 74 Pmax Feed-Only → no direct control

    # ── Strategy — Paid Search extended ──────────────────────────────────────
    # 75 TM Products → no direct control
    # 76 Seasonal → no direct control
    # 77 New/Returning → no direct control
    78: ("google_strategy",  "S013"),  # Match Types
    # 79 Competitor → no direct control
    80: ("google_strategy",  "S010"),  # Brand Campaigns (Search) → same as TM
    81: ("google_strategy",  "S016"),  # GEO-Targeting

    # ── Strategy — Demand Gen/Display ────────────────────────────────────────
    82: ("google_strategy", "S019"),   # Demand Gen - Prospecting
    83: ("google_strategy", "S020"),   # Demand Gen - Remarketing
    # 84 Lookalike Audiences → no direct control
    # 85 AI Generated Images → no direct control
}

# ── KPI row → (source, numeric key) for Account Health section ──────────────
# These are informational cells — populated from the context / date-range KPIs.
# We write to col D only for these rows too.

def _get_kpi_values(ctx: GoogleContext) -> dict:
    """Return {row: value} for the KPI block (rows 3-13)."""
    out = {}
    df02 = get_sheet(ctx, "DATE_RANGE_KPIS")
    df03 = get_sheet(ctx, "YEARLY_KPIS")
    df22 = get_sheet(ctx, "CLIENT_SUCCESS")

    if not df02.empty:
        spend_col = find_col(df02, ["AdSpend"])
        sales_col = find_col(df02, ["AdSales"])
        prev_spend = find_col(df02, ["Prev_AdSpend"])
        prev_sales = find_col(df02, ["Prev_AdSales"])

        r = df02.iloc[0]
        spend = to_float(r[spend_col]) if spend_col else None
        sales = to_float(r[sales_col]) if sales_col else None
        p_spend = to_float(r[prev_spend]) if prev_spend else None
        p_sales = to_float(r[prev_sales]) if prev_sales else None

        if spend and sales and spend > 0:
            out[4] = round(sales / spend, 2)   # Actual ROAS (row 4)
        if spend and p_spend and p_spend > 0:
            out[13] = round((spend - p_spend) / p_spend, 4)  # Spend % change
        if sales and p_sales and p_sales > 0:
            out[9] = round((sales - p_sales) / p_sales, 4)   # Revenue % change

    if not df22.empty:
        row22 = df22.iloc[0]
        budget = to_float(row22.iloc[35]) if len(row22) > 35 else None
        acos_target = to_float(row22.iloc[75]) if len(row22) > 75 else None
        if acos_target and acos_target > 0:
            out[3] = round(1.0 / acos_target, 2)   # ROAS Goal (row 3)
        if budget:
            out[6] = budget                          # Budget Goal

    return out


def _status_fill(val) -> PatternFill:
    if val is True:
        return PatternFill("solid", fgColor="C6EFCE")   # green
    if val is False:
        return PatternFill("solid", fgColor="FFC7CE")   # red
    if val == "PARTIAL":
        return PatternFill("solid", fgColor="FFEB9C")   # amber
    return PatternFill(fill_type=None)


def _safe_cell(ws, row: int, col: int):
    """
    Return the writable top-left cell for a given (row, col), resolving
    merged cell ranges transparently. Returns None if the cell cannot be
    written (e.g. it's inside a merge but has no anchor — shouldn't happen).
    """
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            return ws.cell(merge_range.min_row, merge_range.min_col)
    return None  # shouldn't reach here


def _write(ws, row: int, col: int, value, fill=None, font=None, alignment=None):
    """Write value + optional formatting to a cell, skipping MergedCell non-anchors."""
    cell = _safe_cell(ws, row, col)
    if cell is None:
        return
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment


def write_scorecard_output(
    template_path: str,
    output_path: str,
    all_results: Dict[str, Dict[str, ControlResult]],
    ctx: GoogleContext,
) -> None:
    """
    all_results = {
        "google_health":         {H001: ControlResult, ...},
        "google_mastery":        {M001: ControlResult, ...},
        "google_framework":      {F001: ControlResult, ...},
        "google_strategy":       {S001: ControlResult, ...},
        "google_implementation": {I001: ControlResult, ...},
    }
    """
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # ── Account header ────────────────────────────────────────────────────────
    _write(ws, 1, 1, f"Scorecard — {ctx.hash_name}")
    if ctx.window_start and ctx.window_end:
        _write(ws, 1, 6, f"{ctx.window_start} to {ctx.window_end}")

    # ── KPI block (rows 3-13, col D) ─────────────────────────────────────────
    kpis = _get_kpi_values(ctx)
    for row, val in kpis.items():
        _write(ws, row, 4, val)

    # ── Main checkbox rows ────────────────────────────────────────────────────
    for row_num, (pillar_key, control_id) in ROW_MAP.items():
        pillar_results = all_results.get(pillar_key, {})
        res = pillar_results.get(control_id)
        if res is None:
            continue
        val = _val(res)
        _write(ws, row_num, 4, val, fill=_status_fill(val))

        # Write What We Saw into col F only if the cell is currently empty
        note_cell = _safe_cell(ws, row_num, 6)
        if note_cell is not None and not note_cell.value:
            _write(
                ws, row_num, 6,
                res.what[:200] if res.what else "",
                font=Font(size=8, color="595959"),
                alignment=Alignment(wrap_text=True),
            )

    wb.save(output_path)
    print(f"[writer_scorecard] Saved: {output_path}")
