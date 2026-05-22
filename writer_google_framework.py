"""
writer_google_framework.py
"""
from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from reader_databricks_google import GoogleContext, clean_text
from config_google_framework import SCORING_EXCLUDED, PRIORITY_POINTS, IMPORTANCE
from config import STATUS_OK, STATUS_FLAG, STATUS_PARTIAL

def _compute_score(results, scoring_excluded, importance, priority_points):
    score = 100
    for cid, res in results.items():
        if cid in scoring_excluded:
            continue
        if res.status in (STATUS_FLAG, STATUS_PARTIAL):
            imp = importance.get(cid, 5)
            pts = priority_points.get(imp, -5)
            if res.status == STATUS_PARTIAL:
                pts = pts // 2
            score += pts
    score = max(0, min(100, score))
    grade = "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 60 else "D" if score >= 45 else "F"
    return score, grade

def write_framework_output(template_path, output_path, results, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb["Framework_Analysis"]
    ws_ref  = wb["Framework_Reference"]
    score, grade = _compute_score(results, SCORING_EXCLUDED, IMPORTANCE, PRIORITY_POINTS)
    ws_main["A1"] = f"{ctx.hash_name} — Google Framework Analysis"
    ws_main["B3"] = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}"
    if ctx.window_start and ctx.window_end and ctx.window_days:
        ws_main["B4"] = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
    if ctx.downloaded:
        ws_main["B5"] = ctx.downloaded
        ws_main["B5"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws_main["C8"] = score
    ws_main["D8"] = grade
    cid_to_row = {}
    for r in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f"B{r}"].value).upper()
        if cid:
            cid_to_row[cid] = r
    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f"[writer_framework] WARNING: {cid} not in reference tab — skipping.")
            continue
        rr = cid_to_row[cid]
        ws_ref[f"D{rr}"] = res.status
        ws_ref[f"H{rr}"] = res.what
        ws_ref[f"I{rr}"] = res.why
        for cell in [f"H{rr}", f"I{rr}"]:
            ws_ref[cell].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(output_path)
    print(f"[writer_framework] Saved: {output_path} | Score: {score} | Grade: {grade}")
