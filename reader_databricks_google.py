"""
reader_databricks_google.py
Shared reader for all Google CoE pillars.
Loads all 40 tabs from the Google Databricks export.
Header row is always index 5 (Excel row 6) — hard locked.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ── Public helpers ────────────────────────────────────────────────────────────

def to_float(v) -> Optional[float]:
    if pd.isna(v):
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def to_str(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return "N/A"
    return f"{v * 100:.{decimals}f}%"


def money_str(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    return f"${v:,.2f}"


def num_str(v: Optional[float], decimals: int = 0) -> str:
    if v is None:
        return "N/A"
    return f"{v:,.{decimals}f}"


def clean_text(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", " ")


def _parse_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


def _parse_datetime(v) -> Optional[datetime]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


# ── Tab name registry ─────────────────────────────────────────────────────────
# Maps logical key → tab prefix. Reader uses prefix matching so minor
# tab name changes don't break the agent.

TAB_KEYS = {
    "ADVERTISER_NAME":         "01_Advertiser_Name",
    "DATE_RANGE_KPIS":         "02_Date_Range_KPIs",
    "YEARLY_KPIS":             "03_Yearly_KPIs",
    "L24M_MONTHLY":            "04_L24M_Monthly",
    "MONTHLY_SALES_YOY":       "05_Monthly_Sales",
    "CAMPAIGN_REPORT":         "06_Campaign_Report",
    "CHANNEL_TYPE":            "07_Campaigns_by_Channel",
    "PRODUCT_SHOPPING":        "08_Product_Shopping",
    "KEYWORD_REPORT":          "09_Keyword_Report",
    "SEARCH_TERMS":            "11_Search_Terms_Report",
    "SEARCH_CLASSIFIER":       "12_Search_Terms_Classifier",
    "CAMPAIGN_GOLD":           "13_Campaign_Gold",
    "CAMPAIGN_METADATA":       "14_Campaign_Metadata",
    "STRIPE_INFO":             "15_Stripe_and_Account",
    "DEVICE_BREAKDOWN":        "16_Device_Breakdown",
    "PRODUCT_MONTHLY_KPIS":    "17_Product_Monthly",
    "PMAX_CHANNELS":           "18_PMAX_Channels",
    "MULTICHANNEL_PRODUCTS":   "19_MultiChannel_Products",
    "PRICE_COMPETITIVENESS":   "20_Price_Competitiveness",
    "CAMPAIGN_MONTH_CDM":      "21_Campaign_Month",
    "CLIENT_SUCCESS":          "22_Client_Success",
    "CAMPAIGN_PERF_CDM":       "23_Campaign_Performance",
    "PLA_SUMMARY":             "24_PLA_Summary",
    "KPIS_CDM":                "25_KPIs_CDM",
    "LOCATION_PERF":           "26_Location_Performance",
    "AMAZON_PRODUCT":          "27_Amazon_Product",
    "DPL_PERFORMANCE":         "28_DPL_Performance",
    "SEARCH_TERMS_CDM":        "29_Search_Terms_Performance",
    "FEED_PRODUCTS":           "30_Feed_Products",
    "NEGATIVE_KEYWORDS":       "31_Negative_Keywords",
    "ASSET_GROUPS":            "32_Google_Asset_Groups",
    "ASSETS_EXTENSIONS":       "33_Google_Assets_Extensions",
    "ADVERTISER_DETAILS":      "34_Google_Advertiser_Details",
    "CAMPAIGNS_V2_ENRICHED":   "35_Google_Campaigns_V2_Enriched",
    "PRODUCT_GROUPS":          "36_Google_Product_Groups",
    "AD_GROUP_ADS":            "37_Google_Ad_Group_Ads",
    "CAMPAIGN_SETTINGS":       "38_Google_Campaign_Settings",
    "AD_GROUPS":               "39_Google_Ad_Groups",
    "ACCOUNT_LINKS":           "40_Google_Account_Links",
}


# ── Context dataclass ─────────────────────────────────────────────────────────

@dataclass
class GoogleContext:
    workbook_path: str

    # Identity
    hash_name: str
    tenant_id: str
    account_id: str
    downloaded: Optional[datetime]

    # Eval window
    window_start: Optional[date]
    window_end: Optional[date]
    window_days: Optional[int]
    window_str: str

    # All sheets — keyed by logical TAB_KEY
    sheets: Dict[str, pd.DataFrame] = field(default_factory=dict)


# ── Header extraction ─────────────────────────────────────────────────────────

def _extract_header(path: str) -> Tuple[str, str, str, Optional[date], Optional[date], Optional[datetime]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for s in wb.sheetnames:
        if str(s).strip().lower().startswith("01_"):
            sheet = s
            break
    if sheet is None:
        wb.close()
        return "", "", "", None, None, None

    ws = wb[sheet]
    a1 = ws["A1"].value
    hash_name = re.sub(r"\s*-\s*Advertiser_Name\s*$", "", str(a1 or "").strip(), flags=re.IGNORECASE).strip()

    tenant_id = account_id = ""
    start = end = None
    downloaded_dt = None

    for r in range(1, 25):
        cells = [str(ws.cell(r, c).value or "") for c in range(1, 15)]
        line = " ".join(cells).strip()
        low = line.lower()

        if "tenant id" in low and "advertiser id" in low:
            mt = re.search(r"Tenant\s*ID:\s*([0-9a-fA-F-]{8,})", line)
            ma = re.search(r"Advertiser\s*ID:\s*([0-9]{6,})", line)
            if mt:
                tenant_id = mt.group(1).strip()
            if ma:
                account_id = ma.group(1).strip()

        if "date range" in low:
            m = re.search(r"(\d{4}-\d{2}-\d{2})\s*to\s*(\d{4}-\d{2}-\d{2})", line)
            if m:
                start = _parse_date(m.group(1))
                end = _parse_date(m.group(2))

        if "downloaded" in low and downloaded_dt is None:
            m = re.search(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", line)
            if m:
                downloaded_dt = _parse_datetime(m.group(1))

    wb.close()
    return hash_name, tenant_id, account_id, start, end, downloaded_dt


# ── Main loader ───────────────────────────────────────────────────────────────

def load_google_export(path: str) -> GoogleContext:
    """
    Load a Google Databricks export workbook.
    Returns a GoogleContext with all 40 tabs loaded into sheets dict.
    Sheets are keyed by logical TAB_KEY from TAB_KEYS registry.
    Tabs not in the registry are silently skipped.
    Tabs returning NO DATA are stored as empty DataFrames — callers must check len(df) == 0.
    """
    xls = pd.ExcelFile(path, engine="calamine")

    # Build reverse map: prefix → key
    prefix_to_key = {v: k for k, v in TAB_KEYS.items()}

    def _match_key(sheet_name: str) -> Optional[str]:
        for prefix, key in prefix_to_key.items():
            if sheet_name.startswith(prefix):
                return key
        return None

    sheets: Dict[str, pd.DataFrame] = {}

    for sname in xls.sheet_names:
        key = _match_key(sname)
        if key is None:
            continue
        try:
            df = pd.read_excel(xls, sheet_name=sname, header=5)
            # Drop Unnamed columns from merged title rows
            df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed:\s*\d+$", na=False)].copy()
            # Detect NO DATA tabs
            if len(df.columns) == 1 and "NO DATA" in str(df.columns[0]).upper():
                df = pd.DataFrame()
            sheets[key] = df
        except Exception as e:
            print(f"[reader_google] WARNING: could not load tab {sname}: {e}")
            sheets[key] = pd.DataFrame()

    hash_name, tenant_id, account_id, h_start, h_end, downloaded_dt = _extract_header(path)

    if h_start and h_end:
        window_days = (h_end - h_start).days + 1
        window_str = f"{h_start} to {h_end} ({window_days} days)"
    else:
        window_days = None
        window_str = "UNKNOWN WINDOW"

    return GoogleContext(
        workbook_path=path,
        hash_name=hash_name,
        tenant_id=tenant_id,
        account_id=account_id,
        downloaded=downloaded_dt,
        window_start=h_start,
        window_end=h_end,
        window_days=window_days,
        window_str=window_str,
        sheets=sheets,
    )


def get_sheet(ctx: GoogleContext, key: str) -> pd.DataFrame:
    """
    Return a sheet by TAB_KEY. Returns empty DataFrame if missing or no data.
    Never raises — callers must check len(df) == 0.
    """
    return ctx.sheets.get(key, pd.DataFrame())


def find_col(df: pd.DataFrame, candidates: list) -> Optional[str]:
    """
    Case-insensitive, underscore-insensitive column lookup.
    Returns the actual column name or None.
    """
    norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower().replace(" ", "").replace("_", "")
        if key in norm:
            return norm[key]
    return None
